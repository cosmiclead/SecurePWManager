import os
import json
import time
import base64
import uuid
import webview
from cryptography.hazmat.primitives.ciphers.aead import AESGCM
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
from cryptography.hazmat.primitives import hashes
import openpyxl
import msoffcrypto
from msoffcrypto.format.ooxml import OOXMLFile

import sys

# Base directory for static assets (HTML/CSS/JS bundled inside exe)
if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
    ASSET_DIR = sys._MEIPASS
else:
    ASSET_DIR = os.path.dirname(os.path.abspath(__file__))

# Base directory for persistent user database (hidden .data folder next to executable)
if getattr(sys, 'frozen', False):
    EXE_DIR = os.path.dirname(sys.executable)
else:
    EXE_DIR = os.path.dirname(os.path.abspath(__file__))

DATA_DIR = os.path.join(EXE_DIR, '.data')
os.makedirs(DATA_DIR, exist_ok=True)

# Hide the folder on Windows to prevent accidental deletion
if os.name == 'nt':
    import ctypes
    try:
        # FILE_ATTRIBUTE_HIDDEN = 2
        ctypes.windll.kernel32.SetFileAttributesW(DATA_DIR, 2)
    except Exception as e:
        print(f"Error hiding data folder: {e}")

VAULT_PATH = os.path.join(DATA_DIR, 'vault.enc')
SALT_PATH = os.path.join(DATA_DIR, 'vault.salt')
LOCKOUT_PATH = os.path.join(DATA_DIR, 'lockout.json')
DORSE_PATH = os.path.join(DATA_DIR, 'dorse.json')

# Security Constants
PBKDF2_ITERATIONS = 600000  # OWASP recommended for PBKDF2-HMAC-SHA256
MAX_FAILED_ATTEMPTS = 3
INITIAL_LOCKOUT_DURATION = 30  # seconds

DEFAULT_DORSE = {
   "_":"-...-","A":"...",	"a":"....-.",	"0":"--...",
  "!":".--.-.",	"B":"--..",	"b":"-..--",	"1":".....",
  ";":"..--..",	"C":"..",	  "c":"--.--",	"2":"---..",
  "#":".-..--",	"D":"--.-",	"d":"--.-.",	"3":"-----",
  "$":"...-.-",	"E":".-.",	"e":".....-",	"4":"----.",
  "%":".-...",	"F":"-",	  "f":"--..-",	"5":"...--",
  "&":"...-..-","G":".---",	"g":"-.-..",	"6":"-....",
  "'":"-.-.--",	"H":"-...",	"h":"....--",	"7":"..---",
  "(":".-.-.",	"I":"-..-",	"i":"......",	"8":"....-",
  ")":".-..-.",	"J":".--.",	"j":"---.-",	"9":".----",
  "*":"-..-.",	"K":"-.--",	"k":".---.",			
  "+":".----.",	"L":".--",	"l":"..--",			
  "-":".-.--.",	"M":"...-",	"m":".--..",			
  ".":"..-.-.",	"N":"-.-.",	"n":"..-..",			
  "/":"..-...",	"O":".-..",	"o":"---.",			
  ":":"...---",	"P":".",	  "p":".-..-",			
  "^":".-.-..",	"Q":"-.-",	"q":"...-.",			
  "<":"-.--.-",	"R":"..-.",	"r":"...-..",			
  "=":".-.-.-",	"S":"-..",	"s":"..--.",			
  ">":".-.---",	"T":"---",	"t":".--.-",			
  "?":"...--.",	"U":".-",	  "u":"..-.-",			
  "@":"---...",	"V":"-.",	  "v":"----",			
  "{":"-.--.",	"W":"....",	"w":".-.--",			
  "|":"-....-",	"X":"--",	  "x":"-.---",			
  "}":"..-..-",	"Y":"--.",	"y":"-.-.-",			
  "~":"-.-.-.",	"Z":"..-",	"z":".-.-"		
			
}

class VaultManager:
    """
    Handles file storage, key derivation, AES-256-GCM encryption/decryption,
    and failed login lockout persistence.
    """
    def __init__(self):
        self.key = None
        self.passwords = None

    def is_vault_created(self):
        return os.path.exists(VAULT_PATH) and os.path.exists(SALT_PATH)

    def get_lockout_data(self):
        if not os.path.exists(LOCKOUT_PATH):
            return {"failed_attempts": 0, "lockout_until": 0.0, "lockout_duration": INITIAL_LOCKOUT_DURATION}
        try:
            with open(LOCKOUT_PATH, 'r') as f:
                data = json.load(f)
                if "failed_attempts" not in data or "lockout_until" not in data:
                    raise ValueError("Invalid structure")
                if "lockout_duration" not in data:
                    data["lockout_duration"] = INITIAL_LOCKOUT_DURATION
                return data
        except Exception:
            return {"failed_attempts": 0, "lockout_until": 0.0, "lockout_duration": INITIAL_LOCKOUT_DURATION}

    def save_lockout_data(self, data):
        try:
            with open(LOCKOUT_PATH, 'w') as f:
                json.dump(data, f)
        except Exception as e:
            print(f"Error saving lockout data: {e}")

    def check_lockout(self):
        data = self.get_lockout_data()
        now = time.time()
        remaining = data["lockout_until"] - now
        return max(0, int(remaining))

    def record_failed_attempt(self):
        data = self.get_lockout_data()
        data["failed_attempts"] += 1
        
        if data["failed_attempts"] >= MAX_FAILED_ATTEMPTS:
            mult = 2 ** (data["failed_attempts"] - MAX_FAILED_ATTEMPTS)
            duration = data["lockout_duration"] * mult
            duration = min(duration, 3600)
            data["lockout_until"] = time.time() + duration
            print(f"User locked out for {duration} seconds (Attempts: {data['failed_attempts']})")
        
        self.save_lockout_data(data)
        return self.check_lockout()

    def reset_failed_attempts(self):
        data = self.get_lockout_data()
        data["failed_attempts"] = 0
        data["lockout_until"] = 0.0
        data["lockout_duration"] = INITIAL_LOCKOUT_DURATION
        self.save_lockout_data(data)

    def derive_key(self, master_password, salt):
        kdf = PBKDF2HMAC(
            algorithm=hashes.SHA256(),
            length=32,
            salt=salt,
            iterations=PBKDF2_ITERATIONS
        )
        return kdf.derive(master_password.encode('utf-8'))

    def create_vault(self, master_password):
        if self.is_vault_created():
            return False, "Vault already exists."
        
        try:
            salt = os.urandom(16)
            self.key = self.derive_key(master_password, salt)
            self.passwords = []
            
            with open(SALT_PATH, 'w') as f:
                f.write(base64.b64encode(salt).decode('utf-8'))
                
            self.save_vault()
            self.reset_failed_attempts()
            return True, "Vault created successfully."
        except Exception as e:
            self.clear_sensitive_data()
            return False, f"Failed to create vault: {str(e)}"

    def unlock(self, master_password):
        lockout_secs = self.check_lockout()
        if lockout_secs > 0:
            return False, f"Too many failed attempts. Try again in {lockout_secs} seconds.", lockout_secs

        if not self.is_vault_created():
            return False, "Vault files not found.", 0

        try:
            with open(SALT_PATH, 'r') as f:
                salt_b64 = f.read().strip()
                salt = base64.b64decode(salt_b64)
                
            derived_key = self.derive_key(master_password, salt)
            
            with open(VAULT_PATH, 'rb') as f:
                encrypted_data = f.read()
                
            if len(encrypted_data) < 13:
                raise ValueError("Vault file is corrupted or empty.")
                
            nonce = encrypted_data[:12]
            ciphertext = encrypted_data[12:]
            
            aesgcm = AESGCM(derived_key)
            decrypted_bytes = aesgcm.decrypt(nonce, ciphertext, None)
            
            self.passwords = json.loads(decrypted_bytes.decode('utf-8'))
            self.key = derived_key
            self.reset_failed_attempts()
            return True, "Vault unlocked successfully.", 0
            
        except (Exception, ValueError) as e:
            self.clear_sensitive_data()
            rem_lockout = self.record_failed_attempt()
            
            msg = "Incorrect master password."
            if "corrupted" in str(e).lower() or isinstance(e, json.JSONDecodeError):
                msg = "Vault file is corrupted or invalid."
                
            return False, msg, rem_lockout

    def save_vault(self):
        if self.key is None or self.passwords is None:
            raise PermissionError("Vault is not unlocked.")
            
        plain_bytes = json.dumps(self.passwords).encode('utf-8')
        
        aesgcm = AESGCM(self.key)
        nonce = os.urandom(12)
        ciphertext = aesgcm.encrypt(nonce, plain_bytes, None)
        
        with open(VAULT_PATH, 'wb') as f:
            f.write(nonce + ciphertext)

    def change_master_password(self, current_password, new_password):
        if self.key is None:
            return False, "Vault is locked."
            
        with open(SALT_PATH, 'r') as f:
            salt_b64 = f.read().strip()
            salt = base64.b64decode(salt_b64)
            
        current_derived = self.derive_key(current_password, salt)
        if current_derived != self.key:
            return False, "Incorrect current master password."
            
        try:
            new_salt = os.urandom(16)
            new_key = self.derive_key(new_password, new_salt)
            
            with open(SALT_PATH, 'w') as f:
                f.write(base64.b64encode(new_salt).decode('utf-8'))
                
            self.key = new_key
            self.save_vault()
            return True, "Master password changed successfully."
        except Exception as e:
            return False, f"Failed to change master password: {str(e)}"

    def verify_master_password(self, candidate_password):
        if self.key is None:
            return False
        try:
            with open(SALT_PATH, 'r') as f:
                salt_b64 = f.read().strip()
                salt = base64.b64decode(salt_b64)
            derived_key = self.derive_key(candidate_password, salt)
            return derived_key == self.key
        except Exception:
            return False

    def clear_sensitive_data(self):
        self.key = None
        self.passwords = None

class Api:
    """
    Exposes secure backend API methods to JavaScript via PyWebView's js_api.
    """
    def __init__(self, vault_mgr):
        self.vault_mgr = vault_mgr

    def verify_master_password(self, candidate_password):
        if not candidate_password:
            return False
        return self.vault_mgr.verify_master_password(candidate_password)

    def is_vault_created(self):
        return self.vault_mgr.is_vault_created()

    def get_lockout_time(self):
        return self.vault_mgr.check_lockout()

    def unlock(self, master_password):
        if not master_password:
            return {"success": False, "error": "Password cannot be empty.", "lockout_time": 0}
        success, msg, lockout_secs = self.vault_mgr.unlock(master_password)
        return {"success": success, "error": msg if not success else "", "lockout_time": lockout_secs}

    def create_vault(self, master_password):
        if not master_password or len(master_password) < 8:
            return {"success": False, "error": "Master password must be at least 8 characters long."}
        success, msg = self.vault_mgr.create_vault(master_password)
        return {"success": success, "error": msg if not success else ""}

    def is_locked(self):
        return self.vault_mgr.key is None

    def lock_vault(self):
        self.vault_mgr.clear_sensitive_data()
        return True

    def get_passwords(self):
        if self.vault_mgr.key is None or self.vault_mgr.passwords is None:
            return {"success": False, "error": "Vault is locked."}
        return {"success": True, "passwords": self.vault_mgr.passwords}

    def add_password(self, service, username, password):
        if self.vault_mgr.key is None or self.vault_mgr.passwords is None:
            return {"success": False, "error": "Vault is locked."}
            
        service = service.strip()
        username = username.strip()
        password = password.strip()
        if not service or not username or not password:
            return {"success": False, "error": "All fields are required."}
            
        try:
            entry = {
                "id": str(uuid.uuid4()),
                "service": service,
                "username": username,
                "password": password
            }
            self.vault_mgr.passwords.append(entry)
            self.vault_mgr.save_vault()
            return {"success": True, "entry": entry}
        except Exception as e:
            return {"success": False, "error": f"Failed to save password: {str(e)}"}

    def update_password(self, entry_id, service, username, password):
        if self.vault_mgr.key is None or self.vault_mgr.passwords is None:
            return {"success": False, "error": "Vault is locked."}
            
        service = service.strip()
        username = username.strip()
        password = password.strip()
        if not service or not username or not password:
            return {"success": False, "error": "All fields are required."}
            
        try:
            for entry in self.vault_mgr.passwords:
                if entry["id"] == entry_id:
                    entry["service"] = service
                    entry["username"] = username
                    entry["password"] = password
                    self.vault_mgr.save_vault()
                    return {"success": True, "entry": entry}
            return {"success": False, "error": "Entry not found."}
        except Exception as e:
            return {"success": False, "error": f"Failed to update password: {str(e)}"}

    def delete_password(self, entry_id):
        if self.vault_mgr.key is None or self.vault_mgr.passwords is None:
            return {"success": False, "error": "Vault is locked."}
            
        try:
            for i, entry in enumerate(self.vault_mgr.passwords):
                if entry["id"] == entry_id:
                    self.vault_mgr.passwords.pop(i)
                    self.vault_mgr.save_vault()
                    return {"success": True}
            return {"success": False, "error": "Entry not found."}
        except Exception as e:
            return {"success": False, "error": f"Failed to delete password: {str(e)}"}

    def change_master_password(self, current_password, new_password):
        if self.vault_mgr.key is None:
            return {"success": False, "error": "Vault is locked."}
        if not new_password or len(new_password) < 8:
            return {"success": False, "error": "New master password must be at least 8 characters long."}
        success, msg = self.vault_mgr.change_master_password(current_password, new_password)
        return {"success": success, "error": msg if not success else ""}

    # Dorse Configuration API Methods
    def get_dorse_config(self):
        if not os.path.exists(DORSE_PATH):
            return DEFAULT_DORSE
        try:
            with open(DORSE_PATH, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            return DEFAULT_DORSE

    def save_dorse_config(self, new_config):
        if not isinstance(new_config, dict):
            return {"success": False, "error": "Invalid configuration format."}
        
        for key, val in new_config.items():
            if not isinstance(val, str) or not val:
                return {"success": False, "error": f"Invalid code value for key '{key}'."}
            if any(c.isspace() for c in val):
                return {"success": False, "error": f"Invalid character in code for key '{key}'. Spaces and whitespace are not allowed."}
                
        try:
            with open(DORSE_PATH, 'w', encoding='utf-8') as f:
                json.dump(new_config, f, indent=2, ensure_ascii=False)
            return {"success": True}
        except Exception as e:
            return {"success": False, "error": f"Failed to save configuration: {str(e)}"}

    def reset_dorse_config(self):
        try:
            if os.path.exists(DORSE_PATH):
                os.remove(DORSE_PATH)
            return {"success": True, "config": DEFAULT_DORSE}
        except Exception as e:
            return {"success": False, "error": f"Failed to reset: {str(e)}"}

    def export_credentials_excel(self, master_password):
        if self.vault_mgr.key is None or self.vault_mgr.passwords is None:
            return {"success": False, "error": "Vault is locked."}
        
        if not master_password:
            return {"success": False, "error": "Password cannot be empty."}
            
        # Verify the master password
        if not self.vault_mgr.verify_master_password(master_password):
            return {"success": False, "error": "Incorrect master password."}
            
        if not webview.windows:
            return {"success": False, "error": "Active window not found."}
            
        # Open save file dialog on GUI/Main thread
        try:
            file_path = webview.windows[0].create_file_dialog(
                webview.SAVE_DIALOG,
                file_types=('Excel files (*.xlsx)', 'All files (*.*)'),
                save_filename='credentials_backup.xlsx'
            )
        except Exception as e:
            return {"success": False, "error": f"Failed to open save dialog: {str(e)}"}
        
        if not file_path:
            # User cancelled the save dialog
            return {"success": True, "cancelled": True}
            
        if isinstance(file_path, (tuple, list)):
            if len(file_path) == 0:
                return {"success": True, "cancelled": True}
            file_path = file_path[0]
            
        if not file_path:
            return {"success": True, "cancelled": True}
            
        # Create temp file in DATA_DIR
        temp_file_path = os.path.join(DATA_DIR, f"temp_{uuid.uuid4().hex}.xlsx")
        
        try:
            # Create unencrypted workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Credentials"
            
            # Write headers
            ws.append(["Service / Website", "Username / Email", "Password"])
            
            # Write password entries
            for entry in self.vault_mgr.passwords:
                ws.append([entry.get("service", ""), entry.get("username", ""), entry.get("password", "")])
                
            # Set columns auto-width for professional layout
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 15)
                
            # Save the workbook to temporary file
            wb.save(temp_file_path)
            
            # Encrypt using msoffcrypto-tool OOXMLFile
            with open(temp_file_path, "rb") as f:
                office_file = OOXMLFile(f)
                with open(file_path, "wb") as out:
                    office_file.encrypt(master_password, out)
                    
            return {"success": True, "cancelled": False}
        except Exception as e:
            return {"success": False, "error": f"Failed to export credentials: {str(e)}"}
        finally:
            # Clean up temp file
            if os.path.exists(temp_file_path):
                try:
                    os.remove(temp_file_path)
                except Exception as ex:
                    print(f"Error removing temporary export file: {ex}")

    def import_credentials_excel(self, excel_password):
        if self.vault_mgr.key is None or self.vault_mgr.passwords is None:
            return {"success": False, "error": "Vault is locked."}
            
        if not webview.windows:
            return {"success": False, "error": "Active window not found."}
            
        # Open file dialog on GUI/Main thread
        try:
            file_path = webview.windows[0].create_file_dialog(
                webview.OPEN_DIALOG,
                file_types=('Excel files (*.xlsx)', 'All files (*.*)')
            )
        except Exception as e:
            return {"success": False, "error": f"Failed to open dialog: {str(e)}"}
            
        if not file_path:
            return {"success": True, "cancelled": True}
            
        if isinstance(file_path, (tuple, list)):
            if len(file_path) == 0:
                return {"success": True, "cancelled": True}
            file_path = file_path[0]
            
        if not file_path:
            return {"success": True, "cancelled": True}
            
        # Check if the Excel file is encrypted
        try:
            with open(file_path, "rb") as f:
                office_file = msoffcrypto.OfficeFile(f)
                is_enc = office_file.is_encrypted()
        except Exception as e:
            return {"success": False, "error": f"Failed to check file encryption: {str(e)}"}
            
        temp_decrypted_path = None
        load_path = file_path
        
        if is_enc:
            if not excel_password:
                return {"success": False, "error": "The Excel file is password-protected. Please enter the decryption password."}
                
            temp_decrypted_path = os.path.join(DATA_DIR, f"temp_import_{uuid.uuid4().hex}.xlsx")
            try:
                with open(file_path, "rb") as f:
                    office_file = msoffcrypto.OfficeFile(f)
                    office_file.load_key(password=excel_password)
                    with open(temp_decrypted_path, "wb") as out:
                        office_file.decrypt(out)
                load_path = temp_decrypted_path
            except Exception as e:
                if os.path.exists(temp_decrypted_path):
                    try:
                        os.remove(temp_decrypted_path)
                    except Exception:
                        pass
                return {"success": False, "error": f"Failed to decrypt Excel file: Incorrect password or invalid file format. ({str(e)})"}
                
        # Read the Excel sheet
        try:
            wb = openpyxl.load_workbook(load_path, data_only=True)
            ws = wb.active
            
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return {"success": False, "error": "The Excel sheet is empty."}
                
            header = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
            
            # Map headers to indices
            service_idx = -1
            username_idx = -1
            password_idx = -1
            
            for idx, col_name in enumerate(header):
                if "service" in col_name or "website" in col_name:
                    service_idx = idx
                elif "username" in col_name or "email" in col_name:
                    username_idx = idx
                elif "password" in col_name:
                    password_idx = idx
                    
            if service_idx == -1 or username_idx == -1 or password_idx == -1:
                return {
                    "success": False,
                    "error": "Invalid Excel structure. Could not find column headers for 'Service', 'Username', and 'Password'."
                }
                
            imported_count = 0
            for row_cells in rows[1:]:
                # Skip empty rows
                if all(cell is None or str(cell).strip() == "" for cell in row_cells):
                    continue
                    
                service = str(row_cells[service_idx]).strip() if service_idx < len(row_cells) and row_cells[service_idx] is not None else ""
                username = str(row_cells[username_idx]).strip() if username_idx < len(row_cells) and row_cells[username_idx] is not None else ""
                password = str(row_cells[password_idx]).strip() if password_idx < len(row_cells) and row_cells[password_idx] is not None else ""
                
                if not service or not username or not password:
                    continue # Skip incomplete rows
                    
                entry = {
                    "id": str(uuid.uuid4()),
                    "service": service,
                    "username": username,
                    "password": password
                }
                self.vault_mgr.passwords.append(entry)
                imported_count += 1
                
            # Auto save the vault if we imported anything
            if imported_count > 0:
                self.vault_mgr.save_vault()
                
            return {"success": True, "cancelled": False, "count": imported_count}
            
        except Exception as e:
            return {"success": False, "error": f"Failed to parse Excel file: {str(e)}"}
        finally:
            # Always remove the temporary decrypted file
            if temp_decrypted_path and os.path.exists(temp_decrypted_path):
                try:
                    os.remove(temp_decrypted_path)
                except Exception as ex:
                    print(f"Error removing temporary decrypted import file: {ex}")

def auto_compile_exe():
    if getattr(sys, 'frozen', False):
        return
        
    script_dir = os.path.dirname(os.path.abspath(__file__))
    dist_dir = os.path.join(script_dir, 'dist')
    exe_path = os.path.join(dist_dir, 'app.exe')
    
    if not os.path.exists(exe_path):
        print("dist/app.exe not found. Auto-compiling stand-alone executable with PyInstaller...")
        try:
            import subprocess
            cmd = [
                "pyinstaller",
                "--onefile",
                "--windowed",
                "--add-data", f"index.html{os.pathsep}.",
                "--add-data", f"style.css{os.pathsep}.",
                "--add-data", f"script.js{os.pathsep}.",
                "app.py"
            ]
            subprocess.run(cmd, cwd=script_dir, check=True)
            print("Stand-alone executable successfully built at dist/app.exe")
        except Exception as e:
            print(f"Warning: Could not auto-compile dist/app.exe: {e}")

def main():
    auto_compile_exe()
    vault_mgr = VaultManager()
    api = Api(vault_mgr)
    
    index_html_path = os.path.join(ASSET_DIR, 'index.html')
    
    window = webview.create_window(
        title="Secure Password Manager",
        url=index_html_path,
        js_api=api,
        width=950,
        height=720,
        resizable=True
    )
    
    webview.start()

if __name__ == '__main__':
    main()
